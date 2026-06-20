import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import datetime
import os


# ──────────────────────────────────────────────
#  공통 유틸
# ──────────────────────────────────────────────

def _make_output_dir():
    os.makedirs("output", exist_ok=True)

def _safe_filename(name: str) -> str:
    for ch in r'\/:*?"<>|':
        name = name.replace(ch, "_")
    return name

def _build_summary_df(raw_results: list) -> pd.DataFrame:
    """raw_results → 보고서용 요약 DataFrame"""
    rows = []
    for r in raw_results:
        rows.append({
            "진단 항목":  r.get("rule_name", ""),
            "테이블명":   r.get("table", ""),
            "컬럼명":     r.get("column", ""),
            "총 건수":    r.get("total_cnt", 0),
            "오류 건수":  r.get("error_cnt", 0),
            "오류율(%)":  r.get("error_rate", 0.0),
        })
    return pd.DataFrame(rows)


# ──────────────────────────────────────────────
#  엑셀 보고서
# ──────────────────────────────────────────────

class ExcelReportGenerator:

    # 색상 팔레트
    COLOR_HEADER    = "1F4E79"   # 짙은 남색 (헤더 배경)
    COLOR_TITLE_ROW = "2E75B6"   # 중간 파랑 (소제목 행)
    COLOR_ERROR     = "FF4C4C"   # 빨강 (오류)
    COLOR_OK        = "70AD47"   # 녹색 (정상)
    COLOR_LIGHT     = "DEEAF1"   # 연한 파랑 (짝수 행)
    COLOR_WHITE     = "FFFFFF"

    def _border(self):
        thin = Side(style="thin", color="BFBFBF")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _header_cell(self, ws, row, col, value, bg=None, font_size=11, bold=True, align="center"):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor=bg or self.COLOR_HEADER)
        cell.font = Font(name="맑은 고딕", bold=bold, color=self.COLOR_WHITE, size=font_size)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = self._border()
        return cell

    def _data_cell(self, ws, row, col, value, bg="FFFFFF", bold=False, align="center", color="000000"):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(name="맑은 고딕", bold=bold, color=color, size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = self._border()
        return cell

    # ── 시트1: 진단 개요 ──────────────────────────
    def _sheet_overview(self, wb, project_info: dict, summary_df: pd.DataFrame):
        ws = wb.create_sheet("01_진단개요")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 40

        # 로고 삽입 (있을 때만)
        logo_bytes = project_info.get("logo_bytes")
        logo_row_offset = 0
        if logo_bytes:
            try:
                from openpyxl.drawing.image import Image as XLImage
                import io
                img = XLImage(io.BytesIO(logo_bytes))
                img.width  = 80
                img.height = 50
                ws.add_image(img, "A1")
                ws.row_dimensions[1].height = 45
                ws.row_dimensions[2].height = 10
                logo_row_offset = 2
            except Exception:
                logo_row_offset = 0

        # 타이틀
        title_row = 1 + logo_row_offset
        ws.merge_cells(f"A{title_row}:B{title_row}")
        title_cell = ws[f"A{title_row}"]
        title_cell.value = "데이터 품질 진단 결과 보고서"
        title_cell.font = Font(name="맑은 고딕", bold=True, size=18, color=self.COLOR_HEADER)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[title_row].height = 40

        ws.append([])  # 빈 행

        meta = [
            ("기관명",     project_info.get("org_name", "-")),
            ("사업명",     project_info.get("project_name", "-")),
            ("진단 담당자", project_info.get("manager", "-")),
            ("진단 일시",   datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("대상 테이블", project_info.get("table_name", "-")),
            ("총 진단 항목 수", f"{len(summary_df)} 개"),
            ("오류 발생 항목 수", f"{len(summary_df[summary_df['오류 건수'] > 0])} 개"),
            ("총 점검 건수", f"{summary_df['총 건수'].sum():,} 건"),
            ("총 오류 건수", f"{summary_df['오류 건수'].sum():,} 건"),
        ]

        for i, (label, value) in enumerate(meta, start=3 + logo_row_offset):
            self._header_cell(ws, i, 1, label, bg=self.COLOR_TITLE_ROW, font_size=10)
            cell = ws.cell(row=i, column=2, value=value)
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = self._border()
            ws.row_dimensions[i].height = 22

    # ── 시트2: 상세 결과 ──────────────────────────
    def _sheet_detail(self, wb, summary_df: pd.DataFrame):
        ws = wb.create_sheet("02_상세진단결과")
        ws.sheet_view.showGridLines = False

        headers = ["진단 항목", "테이블명", "컬럼명", "총 건수", "오류 건수", "오류율(%)"]
        col_widths = [40, 18, 22, 12, 12, 12]

        # 헤더
        for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
            self._header_cell(ws, 1, ci, h)
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 28

        # 데이터
        for ri, (_, row) in enumerate(summary_df.iterrows(), start=2):
            bg = self.COLOR_LIGHT if ri % 2 == 0 else self.COLOR_WHITE
            error_cnt = row["오류 건수"]
            error_rate = row["오류율(%)"]

            self._data_cell(ws, ri, 1, row["진단 항목"],  bg=bg, align="left")
            self._data_cell(ws, ri, 2, row["테이블명"],   bg=bg)
            self._data_cell(ws, ri, 3, row["컬럼명"],    bg=bg)
            self._data_cell(ws, ri, 4, row["총 건수"],   bg=bg)

            # 오류 건수: 0이면 녹색, 아니면 빨강
            err_color = self.COLOR_ERROR if error_cnt > 0 else self.COLOR_OK
            self._data_cell(ws, ri, 5, error_cnt,   bg=bg, color=err_color, bold=(error_cnt > 0))
            self._data_cell(ws, ri, 6, f"{error_rate:.2f}%", bg=bg, color=err_color, bold=(error_cnt > 0))
            ws.row_dimensions[ri].height = 20

        # 자동 필터
        ws.auto_filter.ref = f"A1:F{1 + len(summary_df)}"

    # ── 시트3: 영역별 요약 ────────────────────────
    def _sheet_by_dimension(self, wb, summary_df: pd.DataFrame):
        ws = wb.create_sheet("03_영역별요약")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14

        dimension_map = {
            "완전성": "완전성 (Completeness)",
            "일관성": "일관성 (Consistency)",
            "정확성": "정확성 (Accuracy)",
            "유용성": "유용성 (Usefulness)",
            "유일성": "유일성 (Uniqueness)",
            "유효성": "유효성 (Validity)",
        }

        self._header_cell(ws, 1, 1, "품질 영역",    font_size=11)
        self._header_cell(ws, 1, 2, "진단 항목 수", font_size=11)
        self._header_cell(ws, 1, 3, "오류 항목 수", font_size=11)
        self._header_cell(ws, 1, 4, "평균 오류율",  font_size=11)
        ws.row_dimensions[1].height = 26

        for ri, (keyword, label) in enumerate(dimension_map.items(), start=2):
            mask = summary_df["진단 항목"].str.contains(keyword, na=False)
            sub = summary_df[mask]
            total_items = len(sub)
            error_items = len(sub[sub["오류 건수"] > 0])
            avg_rate    = sub["오류율(%)"].mean() if total_items > 0 else 0.0
            bg = self.COLOR_LIGHT if ri % 2 == 0 else self.COLOR_WHITE

            self._data_cell(ws, ri, 1, label,              bg=bg, align="left")
            self._data_cell(ws, ri, 2, total_items,        bg=bg)
            self._data_cell(ws, ri, 3, error_items,        bg=bg,
                            color=self.COLOR_ERROR if error_items > 0 else self.COLOR_OK,
                            bold=(error_items > 0))
            self._data_cell(ws, ri, 4, f"{avg_rate:.2f}%", bg=bg)
            ws.row_dimensions[ri].height = 22

        # 차트
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = "품질 영역별 오류 현황"
        chart.y_axis.title = "건수"
        chart.x_axis.title = "영역"
        chart.width = 16
        chart.height = 10

        data   = Reference(ws, min_col=3, min_row=1, max_row=7)
        cats   = Reference(ws, min_col=1, min_row=2, max_row=7)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "F2")

    # ── 퍼블릭 메서드 ────────────────────────────
    def generate(self, project_info: dict, raw_results: list, output_path: str) -> str:
        _make_output_dir()
        summary_df = _build_summary_df(raw_results)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # 기본 시트 제거

        self._sheet_overview(wb, project_info, summary_df)
        self._sheet_detail(wb, summary_df)
        self._sheet_by_dimension(wb, summary_df)

        wb.save(output_path)
        return output_path


# ──────────────────────────────────────────────
#  PDF 보고서
# ──────────────────────────────────────────────

class PDFReportGenerator:

    def _register_font(self):
        """한글 폰트 등록 (나눔고딕 or 기본 폰트 fallback)"""
        font_paths = [
            "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
            "/usr/share/fonts/nanum/NanumGothic.ttf",
        ]
        for fp in font_paths:
            if os.path.exists(fp):
                pdfmetrics.registerFont(TTFont("NanumGothic", fp))
                return "NanumGothic"
        return "Helvetica"  # fallback

    def _styles(self, font_name: str):
        base = getSampleStyleSheet()
        custom = {
            "title": ParagraphStyle("ReportTitle",
                fontName=font_name, fontSize=20, textColor=colors.HexColor("#1F4E79"),
                alignment=TA_CENTER, spaceAfter=8, leading=28),
            "heading": ParagraphStyle("SectionHeading",
                fontName=font_name, fontSize=13, textColor=colors.HexColor("#2E75B6"),
                spaceBefore=14, spaceAfter=6, leading=18),
            "body": ParagraphStyle("BodyText",
                fontName=font_name, fontSize=9, textColor=colors.black,
                alignment=TA_LEFT, leading=14, spaceAfter=4),
            "small": ParagraphStyle("SmallText",
                fontName=font_name, fontSize=8, textColor=colors.grey,
                alignment=TA_CENTER, leading=12),
        }
        return custom

    def _meta_table(self, project_info: dict, summary_df: pd.DataFrame, font_name: str):
        data = [
            ["항목", "내용"],
            ["기관명",         project_info.get("org_name", "-")],
            ["사업명",         project_info.get("project_name", "-")],
            ["진단 담당자",     project_info.get("manager", "-")],
            ["진단 일시",       datetime.datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["대상 테이블",     project_info.get("table_name", "-")],
            ["총 진단 항목",    f"{len(summary_df)} 개"],
            ["오류 발생 항목",  f"{len(summary_df[summary_df['오류 건수'] > 0])} 개"],
            ["총 점검 건수",    f"{summary_df['총 건수'].sum():,} 건"],
            ["총 오류 건수",    f"{summary_df['오류 건수'].sum():,} 건"],
        ]
        t = Table(data, colWidths=[50*mm, 110*mm])
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#1F4E79")),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
            ("BACKGROUND",  (0, 1), (0, -1),  colors.HexColor("#DEEAF1")),
            ("FONTNAME",    (0, 0), (-1, -1), font_name),
            ("FONTSIZE",    (0, 0), (-1, -1), 9),
            ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F5F9FF"), colors.white]),
            ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
            ("TOPPADDING",  (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",(0,0), (-1, -1), 5),
        ]))
        return t

    def _detail_table(self, summary_df: pd.DataFrame, font_name: str):
        headers = ["진단 항목", "컬럼명", "총 건수", "오류 건수", "오류율(%)"]
        data = [headers]
        for _, row in summary_df.iterrows():
            data.append([
                row["진단 항목"],
                row["컬럼명"],
                f"{int(row['총 건수']):,}",
                f"{int(row['오류 건수']):,}",
                f"{row['오류율(%)']:.2f}%",
            ])

        col_w = [75*mm, 35*mm, 22*mm, 22*mm, 22*mm]
        t = Table(data, colWidths=col_w, repeatRows=1)

        style = [
            ("BACKGROUND",   (0, 0), (-1, 0),  colors.HexColor("#1F4E79")),
            ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",     (0, 0), (-1, -1), font_name),
            ("FONTSIZE",     (0, 0), (-1, -1), 8),
            ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
            ("ALIGN",        (0, 1), (0, -1),  "LEFT"),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ("GRID",         (0, 0), (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
            ("TOPPADDING",   (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS",(0,1), (-1, -1), [colors.HexColor("#F5F9FF"), colors.white]),
        ]
        # 오류 건수 > 0 행 → 빨강 표시
        for i, (_, row) in enumerate(summary_df.iterrows(), start=1):
            if row["오류 건수"] > 0:
                style.append(("TEXTCOLOR", (3, i), (4, i), colors.HexColor("#CC0000")))
                style.append(("FONTNAME",  (3, i), (4, i), font_name))

        t.setStyle(TableStyle(style))
        return t

    def generate(self, project_info: dict, raw_results: list, output_path: str) -> str:
        _make_output_dir()
        summary_df = _build_summary_df(raw_results)
        font_name  = self._register_font()
        styles     = self._styles(font_name)

        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=20*mm, leftMargin=20*mm,
            topMargin=20*mm,   bottomMargin=20*mm,
        )

        story = []

        # ── 표지 ──
        story.append(Spacer(1, 20*mm))

        # 로고 삽입 (있을 때만)
        logo_bytes = project_info.get("logo_bytes")
        if logo_bytes:
            try:
                from reportlab.platypus import Image as RLImage
                import io
                logo_img = RLImage(io.BytesIO(logo_bytes), width=60*mm, height=40*mm,
                                   kind='proportional')
                logo_img.hAlign = 'CENTER'
                story.append(logo_img)
                story.append(Spacer(1, 8*mm))
            except Exception:
                pass

        story.append(Paragraph("데이터 품질 진단 결과 보고서", styles["title"]))
        story.append(Spacer(1, 6*mm))
        story.append(Paragraph(
            f"{project_info.get('org_name', '')} | {project_info.get('project_name', '')}",
            styles["small"]
        ))
        story.append(Spacer(1, 16*mm))

        # ── 진단 개요 ──
        story.append(Paragraph("1. 진단 개요", styles["heading"]))
        story.append(self._meta_table(project_info, summary_df, font_name))
        story.append(Spacer(1, 8*mm))

        # ── 상세 결과 ──
        story.append(Paragraph("2. 상세 진단 결과", styles["heading"]))
        story.append(self._detail_table(summary_df, font_name))
        story.append(Spacer(1, 8*mm))

        # ── 종합 의견 ──
        story.append(Paragraph("3. 종합 의견", styles["heading"]))
        total_error = int(summary_df["오류 건수"].sum())
        total_items = len(summary_df)
        error_items = len(summary_df[summary_df["오류 건수"] > 0])
        opinion = (
            f"총 {total_items}개 진단 항목 중 {error_items}개 항목에서 오류가 발생하였으며, "
            f"총 {total_error:,}건의 오류 데이터가 확인되었습니다. "
            "상세 오류 내역은 별첨 엑셀 파일을 참조하시기 바랍니다."
        )
        story.append(Paragraph(opinion, styles["body"]))

        doc.build(story)
        return output_path


# ──────────────────────────────────────────────
#  통합 인터페이스
# ──────────────────────────────────────────────

class ReportGenerator:
    """Streamlit 6_report.py에서 호출하는 단일 진입점"""

    def generate_excel(self, project_info: dict, raw_results: list) -> str:
        fname = _safe_filename(project_info.get("project_name", "진단결과"))
        path  = f"output/{fname}_품질진단결과.xlsx"
        return ExcelReportGenerator().generate(project_info, raw_results, path)

    def generate_pdf(self, project_info: dict, raw_results: list) -> str:
        fname = _safe_filename(project_info.get("project_name", "진단결과"))
        path  = f"output/{fname}_품질진단결과.pdf"
        return PDFReportGenerator().generate(project_info, raw_results, path)